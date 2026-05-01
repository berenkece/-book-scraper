import csv
import openpyxl

with open("books.csv", "r", encoding = "utf-8") as f:
    reader = csv.DictReader(f)
    books = list(reader)

workbook = openpyxl.Workbook()

worksheet1 = workbook.active
worksheet1.title = "Books"
worksheet1.append(["Title", "Price", "Rating", "Stock"])
for book in books:
    worksheet1.append([book["title"], book["price"], book["rating"], book["stock"]])

worksheet2 = workbook.create_sheet(title="Summary")
worksheet2.append(["Title", "Price","Rating", "Stock"])
for book in books:
    if book["rating"] == "Five":
        worksheet2.append([book["title"], book["price"], book["rating"], book["stock"]])    

worksheet3 = workbook.create_sheet("Alphabetical")
worksheet3.append(["Title", "Price", "Rating", "Stock"])
for book in sorted(books, key=lambda x: x["title"]):
    worksheet3.append([book["title"], book["price"], book["rating"], book["stock"]])

workbook.save("books_analysis.xlsx")
print("Data analysis saved to books_analysis.xlsx")